# -*- coding: utf-8 -*-
"""Generate Smart Institute AI Implementation Tracker (xlsx + csv)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os

OUT_DIR = r"c:\Users\ACER\Smart Institute AI"
COLS = [
    "Section",
    "Sub-Section",
    "Item / Task",
    "Layer",
    "Detail / Expected Result",
    "Status",
    "Assigned Dev",
    "Priority",
    "Sprint / Phase",
    "Notes",
]

# Status default for all build tasks
TODO = "Not Started"

def t(section, sub, item, layer, detail, assignee, priority, phase, notes=""):
    return [section, sub, item, layer, detail, TODO, assignee, priority, phase, notes]


# ---------------------------------------------------------------------------
# SHEET DATA
# ---------------------------------------------------------------------------

overview_header_rows = [
    ["Smart Institute AI — Implementation Tracker", "", "", "", "", "", "", "", "", ""],
    [
        "Source: Client Process Note (WTI Institution) + Smart Institution ERP & AI Functional Spec (Precious / approved by Heramb) | MERN | UI parity with Celeris ERP WTI (Dashboard / Masters / Transactions / Reports)",
        "Arnav + Precious",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ],
    ["", "", "", "", "", "", "", "", "", ""],
    ["Sheet", "Scope", "Item Count", "Primary Owner", "", "", "", "", "", ""],
]

overview_sheets_index = [
    ["01 Architecture-Foundation", "MERN scaffold, WTI-style ERP shell, merge-safe routing, shared UI kit, seeds, audit log", 22, "Arnav (scaffold) + shared routing contract"],
    ["02 RBAC-Auth-Users", "Roles, permissions, login, user management, permission-driven left nav (7 modules)", 18, "Arnav"],
    ["03 Shared-Platform-Masters", "Academic Year, Departments, Programs/Courses, Institution settings (foundation both consume)", 14, "Arnav owns models/APIs; Precious consumes"],
    ["04 Admissions", "Full admissions lifecycle: Masters + Transactions + Reports + AI OCR stub", 28, "Arnav"],
    ["05 Students", "Student lifecycle, hostel/transport, promotion, alumni", 24, "Arnav"],
    ["06 Faculty", "Faculty profiles, subject allocation, leave, performance", 22, "Precious"],
    ["07 Academics", "Curriculum, batches, timetable (AI-assist), attendance, internal assessments", 26, "Precious"],
    ["08 LMS", "Courses, content, assignments, quizzes, live class stub, progress", 22, "Precious"],
    ["09 Exams", "Scheduling, hall tickets, marks, results, transcripts, certificates", 24, "Precious"],
    ["10 Fees", "Fee structure, collection, receipts, scholarships, refunds, outstanding", 26, "Arnav"],
    ["11 Dashboard-Reports", "Institution dashboard + cross-module MIS; module reports owned with parent module", 16, "Split — see sheet"],
    ["12 Phases-Acceptance", "Phase 0–3 milestones, UI parity checklist vs erp.smart-aiapps.com/app/wti, UAT", 24, "Both"],
]

merge_rules = [
    ["MERGE-CONFLICT RULES (READ BEFORE CODING)", "", "", "", "", "", "", "", "", ""],
    ["1. Arnav owns backend/routes/coreRoutes.js + frontend/src/routes/coreRoutes.jsx (Admissions, Students, Fees, Dashboard, Auth, Users, Shared Masters). Precious owns academicRoutes equivalents (Faculty, Academics, LMS, Exams). index only imports both aggregators.", "", "", "", "", "", "", "", "", ""],
    ["2. Never edit the other dev's route aggregator; register new routes only in your own file (coreRoutes vs academicRoutes).", "", "", "", "", "", "", "", "", ""],
    ["3. Arnav owns Layout.jsx, Sidebar.jsx, TopBar.jsx, theme tokens (Celeris/WTI look). Precious MUST NOT edit shell — add menu items via permission seeds + nav config keys only.", "", "", "", "", "", "", "", "", ""],
    ["4. Shared components have a single owner (Notes column). Import — do not duplicate DataTable, PageHeader, MasterCardGrid, StatusBadge, etc.", "", "", "", "", "", "", "", "", ""],
    ["5. Shared foundation masters (AcademicYear, Department, Program/Course) = Arnav. Precious consumes via API only — never edit those model files.", "", "", "", "", "", "", "", "", ""],
    ["6. Student model + enrollment APIs = Arnav. Faculty/Academics/LMS/Exams read Student via API; do not fork Student schema.", "", "", "", "", "", "", "", "", ""],
    ["7. Subject Master + Semester + Classroom + Batch = Precious (Academics). Admissions maps Program only; subject-level work stays with Precious.", "", "", "", "", "", "", "", "", ""],
    ["8. Fee Structure / collection / receipt = Arnav. Admissions fee-payment confirmation calls Fees APIs owned by Arnav.", "", "", "", "", "", "", "", "", ""],
    ["9. Phase 0 scaffold + shell + RBAC merged by Arnav before Precious starts feature branches.", "", "", "", "", "", "", "", "", ""],
    ["10. Cross-module contracts (admission→student enrollment, fee.paid, exam.resultPublished, semester.promote) documented in docs/contracts.md — agree before coding.", "", "", "", "", "", "", "", "", ""],
    ["11. UI parity: match Celeris ERP WTI patterns — left module nav, module tabs Dashboard|Masters|Transactions|Reports, master card grid, blue banner module header. Reference: https://erp.smart-aiapps.com/app/wti/", "", "", "", "", "", "", "", "", ""],
    ["12. Folder ownership: frontend/src/pages/admissions|students|fees|dashboard + api/core/ = Arnav. pages/faculty|academics|lms|exams + api/academic/ = Precious.", "", "", "", "", "", "", "", "", ""],
]

# ---- 01 Architecture ----
sheet_01 = [
    t("1. TECHNOLOGY STACK", "", "", "", "", "", "", "", ""),
    t("Stack", "Monorepo", "Project scaffold (backend/ + frontend/)", "All",
      "Root scripts; backend Node+Express .js; frontend React+Vite .jsx; .env.example for MONGO_URI, JWT_SECRET, PORT, CLIENT_URL.",
      "Arnav", "Critical", "Phase 0", "Precious clones after scaffold merged — do not scaffold in parallel."),
    t("Stack", "Frontend", "React 18 + Vite (frontend/)", "Frontend",
      "Responsive app; path aliases; ESLint. Mobile-friendly breakpoints for ERP tables.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Stack", "Backend", "Node.js + Express API (backend/)", "Backend",
      "REST folders: models/, routes/, controllers/, services/, middleware/, seeds/, utils/, config/.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Stack", "Database", "MongoDB + Mongoose connection", "DB",
      "config/db.js; GET /api/health.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Stack", "Auth", "JWT login + /api/auth/me", "Backend",
      "POST /api/auth/login; Bearer token; password_hash; /me returns user + permissions + institution context.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Stack", "Charts", "Recharts (or Chart.js) shared wrappers", "Frontend",
      "frontend/src/components/charts/ for dashboard + MIS reuse.",
      "Arnav", "Medium", "Phase 2", "Precious imports wrappers — does not rewrite chart kit."),
    t("Stack", "Email", "Email service stub (invoice/receipt later)", "Backend",
      "utils/mailer.js stub; env SMTP placeholders. Wired in Fees Phase 2 for Tax Invoice + Receipt auto-email (client process note).",
      "Arnav", "High", "Phase 2", "Aligns with client automated customer communication."),
    t("2. FOLDER STRUCTURE & MERGE-SAFE ROUTING", "", "", "", "", "", "", "", ""),
    t("Architecture", "Route split", "backend/routes/coreRoutes.js", "Backend",
      "Arnav mounts Admissions, Students, Fees, Dashboard, Auth, Users, Shared Masters here.",
      "Arnav", "Critical", "Phase 0", "Avoids both editing routes/index.js."),
    t("Architecture", "Route split", "backend/routes/academicRoutes.js", "Backend",
      "Precious mounts Faculty, Academics, LMS, Exams routes here.",
      "Precious", "Critical", "Phase 0", ""),
    t("Architecture", "Route split", "frontend/src/routes/coreRoutes.jsx", "Frontend",
      "Lazy routes: auth, dashboard, admissions, students, fees, users, shared masters screens.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Architecture", "Route split", "frontend/src/routes/academicRoutes.jsx", "Frontend",
      "Lazy routes: faculty, academics, lms, exams.",
      "Precious", "Critical", "Phase 0", ""),
    t("Architecture", "App shell", "WTI-style ERP Layout + Sidebar + TopBar", "Frontend",
      "Match Celeris WTI: collapsible left nav with 7 modules (Admissions, Students, Faculty, Academics, LMS, Exams, Fees); top context bar (workspace, search, FY, location); blue module banner.",
      "Arnav", "Critical", "Phase 0", "Precious does not edit Layout/Sidebar/TopBar."),
    t("Architecture", "Module chrome", "ModuleShell — Dashboard | Masters | Transactions | Reports tabs", "Frontend",
      "Shared ModuleShell.jsx owned by Arnav; each module page composes it. Visual parity with /app/wti/masters card layout.",
      "Arnav", "Critical", "Phase 0", "Precious imports ModuleShell only."),
    t("Architecture", "UI kit", "Shared components: MasterCardGrid, DataTable, PageHeader, StatusBadge, FormDrawer", "Frontend",
      "frontend/src/components/ui/ — Celeris-like cards for Masters landing; tables for lists; consistent primary blue accents.",
      "Arnav", "Critical", "Phase 0", "Single owner; Precious imports."),
    t("Architecture", "API client", "api/client.js + api/core/ + api/academic/", "Frontend",
      "Axios + auth interceptor; split API folders by owner (core = Arnav, academic = Precious).",
      "Arnav", "High", "Phase 0", "Precious creates only api/academic/* files."),
    t("Architecture", "Institution context", "models/Institution.js + FY/Location context", "DB",
      "Seed institution; financial year + location in TopBar (parity with ERP context bar).",
      "Arnav", "High", "Phase 0", ""),
    t("Architecture", "Error handling", "Global error middleware + response envelope", "Backend",
      "{ success, data, message } shape; 4xx/5xx handlers.",
      "Arnav", "Medium", "Phase 0", ""),
    t("Architecture", "Audit log", "models/AuditLog.js + helper", "Backend",
      "Log admission approvals, fee overrides, result publish, permission changes, refunds.",
      "Arnav", "High", "Phase 0", "Precious calls writeAudit helper — does not edit model."),
    t("Architecture", "Contracts doc", "docs/contracts.md cross-module events", "Docs",
      "Document: admission.enrolled, fee.paid, result.published, student.promoted, student.alumni.",
      "Both", "Critical", "Phase 0", "Agree before feature coding."),
    t("Architecture", "Seed runner", "backend/seeds/index.js", "Backend",
      "Seed admin user, roles, permissions, institution, sample academic year.",
      "Arnav", "Critical", "Phase 0", ""),
]

# ---- 02 RBAC ----
sheet_02 = [
    t("STEP ZERO — ROLES & PERMISSIONS (BEFORE FEATURE MODULES)", "", "", "", "", "", "", "", ""),
    t("RBAC", "DB", "models/Role.js", "DB",
      "Seed roles: Super Admin, Institution Admin, Admission Officer, Faculty, Exam Controller, Accountant, Student (portal later).",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "DB", "models/Permission.js", "DB",
      "module + action (view|create|edit|delete|approve). Modules: admissions, students, faculty, academics, lms, exams, fees, dashboard, users, settings.",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "DB", "models/RolePermission.js", "DB",
      "M2M role↔permission; drives dynamic left nav.",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "DB", "models/UserMenuOverride.js", "DB",
      "Per-user grant/deny overrides.",
      "Arnav", "High", "Phase 0", ""),
    t("RBAC", "DB", "models/User.js", "DB",
      "name, email, phone, password_hash, role_id, department_id?, is_active, linked_faculty_id?, linked_student_id?.",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "Backend", "middleware/requirePermission.js", "Backend",
      "requirePermission(module, action); resolves RolePermission + overrides.",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "Backend", "GET /api/auth/me permissions payload", "Backend",
      "Returns permission set for Sidebar + route guards.",
      "Arnav", "Critical", "Phase 0", ""),
    t("RBAC", "Frontend", "PermissionContext + usePermission + ProtectedRoute", "Frontend",
      "No scattered role=== checks; menu built from permissions.",
      "Arnav", "Critical", "Phase 0", ""),
    t("USER MANAGEMENT", "", "", "", "", "", "", "", ""),
    t("User Mgmt", "Backend", "CRUD /api/users", "Backend",
      "Create/edit/soft-deactivate; never hard-delete users with history.",
      "Arnav", "Critical", "Phase 0", ""),
    t("User Mgmt", "Backend", "PUT /api/users/:id/permission-overrides", "Backend",
      "Admin toggles overrides.",
      "Arnav", "High", "Phase 1", ""),
    t("User Mgmt", "Frontend", "pages/users/UserList.jsx + UserForm.jsx", "Frontend",
      "Admin-only user management screens inside shell.",
      "Arnav", "High", "Phase 1", ""),
    t("User Mgmt", "Frontend", "pages/auth/Login.jsx", "Frontend",
      "Email + password; post-login redirect to Dashboard or first permitted module.",
      "Arnav", "Critical", "Phase 0", "UI clean/professional — match ERP login simplicity."),
    t("Nav", "Seeds", "Permission seeds for 7 left-nav modules", "DB",
      "Ensure Admissions, Students, Faculty, Academics, LMS, Exams, Fees appear when permitted.",
      "Arnav", "Critical", "Phase 0", "Precious requests new permission keys via seed PR owned by Arnav OR documents keys in contracts for Arnav to add."),
    t("Nav", "Config", "navConfig.js module metadata (icon, path, order)", "Frontend",
      "Single nav metadata file owned by Arnav; Precious only proposes new entries via PR note — Arnav merges nav keys.",
      "Arnav", "Critical", "Phase 0", "Prevents Sidebar edit conflicts."),
    t("Audit", "Backend", "Log permission & user changes", "Backend",
      "Write AuditLog on role/permission/user updates.",
      "Arnav", "Medium", "Phase 1", ""),
]

# ---- 03 Shared masters ----
sheet_03 = [
    t("ARNAV — FOUNDATION MASTERS (PRECIOUS CONSUMES VIA API)", "", "", "", "", "", "", "", ""),
    t("Academic Year", "DB", "models/AcademicYear.js", "DB",
      "name, code, start_date, end_date, is_current, is_active.",
      "Arnav", "Critical", "Phase 1", "Maps to WTI Academic Year master + Admissions master list."),
    t("Academic Year", "Backend", "CRUD /api/academic-years", "Backend",
      "In coreRoutes; set-current endpoint.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Academic Year", "Frontend", "pages/settings/AcademicYearList/Form", "Frontend",
      "Masters-style list under Settings or Shared Masters; card entry from platform masters grid.",
      "Arnav", "High", "Phase 1", ""),
    t("Department", "DB", "models/Department.js", "DB",
      "name, code, head_user_id?, is_active.",
      "Arnav", "Critical", "Phase 1", "WTI Department Master parity."),
    t("Department", "Backend", "CRUD /api/departments", "Backend",
      "coreRoutes.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Department", "Frontend", "pages/settings/DepartmentList/Form", "Frontend",
      "CRUD UI.",
      "Arnav", "High", "Phase 1", ""),
    t("Program/Course", "DB", "models/Program.js (Course Master)", "DB",
      "name, code, department_id, duration, program_type (trade/diploma/degree), intake_default, is_active. Description field for WTI Course Master card.",
      "Arnav", "Critical", "Phase 1", "WTI Course Master; Precious links subjects later via program_id."),
    t("Program/Course", "Backend", "CRUD /api/programs", "Backend",
      "coreRoutes.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Program/Course", "Frontend", "pages/settings/ProgramList/Form", "Frontend",
      "Course master UI matching WTI card description: trade/diploma courses.",
      "Arnav", "High", "Phase 1", ""),
    t("Institution Settings", "DB/UI", "Institution profile + FY display settings", "All",
      "Name, address, logo stub, default FY — TopBar context.",
      "Arnav", "Medium", "Phase 1", ""),
    t("Shared Masters Hub", "Frontend", "pages/settings/SharedMastersHub.jsx", "Frontend",
      "Card grid (WTI Masters style) linking Academic Year, Department, Program — 11-module visual language.",
      "Arnav", "High", "Phase 1", "UI reference: /app/wti/masters."),
    t("API contracts", "Docs", "Document GET list shapes for Precious consumers", "Docs",
      "programs, departments, academic-years list DTOs in docs/contracts.md.",
      "Arnav", "High", "Phase 1", "Precious blocked until these APIs stable."),
    t("Seeds", "DB", "Seed sample AY, Depts, Programs", "DB",
      "Enough data for Faculty/Academics development.",
      "Arnav", "High", "Phase 1", ""),
]

# ---- 04 Admissions ----
sheet_04 = [
    t("MODULE SHELL — ADMISSIONS", "", "", "", "", "", "", "", ""),
    t("Admissions", "Frontend", "Admissions module routes + ModuleShell tabs", "Frontend",
      "Dashboard | Masters | Transactions | Reports under /admissions/* — same pattern as WTI.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Admissions", "Frontend", "Admissions Masters card grid", "Frontend",
      "Cards for: Academic Year (link), Admission Cycle, Programs, Departments (link), Intake, Categories, Eligibility, Document Checklist, Fee Mapping, Scholarship Rules, Statuses.",
      "Arnav", "Critical", "Phase 1", "Visual parity with WTI masters cards."),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Admission Cycle", "DB", "models/AdmissionCycle.js", "DB",
      "academic_year_id, name, start_date, end_date, status (draft|open|closed).",
      "Arnav", "Critical", "Phase 1", ""),
    t("Intake Capacity", "DB", "models/IntakeCapacity.js", "DB",
      "cycle_id, program_id, category_id?, seats.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Admission Category", "DB", "models/AdmissionCategory.js", "DB",
      "name, code, quota_percent?, is_active.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Eligibility", "DB", "models/EligibilityCriteria.js", "DB",
      "program_id, min_qualification, min_marks, rules_json.",
      "Arnav", "High", "Phase 1", ""),
    t("Document Checklist", "DB", "models/AdmissionDocumentType.js", "DB",
      "name, mandatory, applies_to_program_ids[].",
      "Arnav", "High", "Phase 1", ""),
    t("Fee Structure Mapping", "DB", "Admission↔FeeStructure link fields", "DB",
      "Map program + cycle to fee_structure_id (Fees module).",
      "Arnav", "Critical", "Phase 1", "Coordinates with Fees masters owned by Arnav."),
    t("Scholarship Rules", "DB", "models/ScholarshipRule.js (admission-facing)", "DB",
      "name, criteria_json, benefit_json; shared with Fees adjustments.",
      "Arnav", "High", "Phase 1", ""),
    t("Admission Status", "DB", "Status enum/master", "DB",
      "Applied, Docs Pending, Eligible, Merit Listed, Allotted, Approved, Fee Pending, Enrolled, Rejected, Waitlisted.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Masters", "Backend/UI", "CRUD APIs + list/form pages for all admission masters", "All",
      "Each master: model, routes in coreRoutes, List+Form pages under pages/admissions/masters/.",
      "Arnav", "Critical", "Phase 1", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Application", "DB", "models/AdmissionApplication.js", "DB",
      "personal, academic, program_preferences[], category_id, cycle_id, status, merit_score?.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Application", "Backend", "POST/GET/PUT /api/admissions/applications", "Backend",
      "Register application; list/filter by cycle/status.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Documents", "Backend", "Document upload + verification endpoints", "Backend",
      "Upload files; verify/reject checklist items; AI OCR stub hook for authenticity/completeness.",
      "Arnav", "High", "Phase 1", "AI: OCR stub — flag Phase 2 for real OCR."),
    t("Eligibility", "Backend", "Eligibility validation service", "Backend",
      "Validate applicant vs program criteria; set Eligible/Ineligible.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Merit", "Backend", "Merit list generation", "Backend",
      "Rank by merit_score + category quotas; generate MeritList records.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Counseling", "Backend", "Seat allocation / counseling", "Backend",
      "Allot seats against IntakeCapacity; waitlist overflow.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Approval", "Backend", "Admission approval action", "Backend",
      "Committee approve/reject allotted seat; audit log.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Fee confirm", "Backend", "Fee payment confirmation → enrollment trigger", "Backend",
      "On fee.paid (Fees API), convert applicant → Student; generate roll_no + student_id.",
      "Arnav", "Critical", "Phase 1", "Contract: admission.enrolled."),
    t("Enrollment", "Backend", "Student enrollment service", "Backend",
      "Create Student from application; no re-entry of personal data.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Transactions UI", "Frontend", "Application list, detail, verify, allot, approve screens", "Frontend",
      "pages/admissions/transactions/* workflow screens.",
      "Arnav", "Critical", "Phase 1", ""),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "Backend/UI", "Admission Summary, Application Status, Seat Availability, Pending, Merit List, Conversion", "All",
      "Filter by cycle/program; export CSV; charts where useful.",
      "Arnav", "High", "Phase 2", "Maps client: Student admission reports."),
    t("Dashboard", "Frontend", "Admissions mini-dashboard widgets", "Frontend",
      "Applications today, pending verify, seats left, conversion %.",
      "Arnav", "Medium", "Phase 2", ""),
]

# ---- 05 Students ----
sheet_05 = [
    t("MODULE SHELL — STUDENTS", "", "", "", "", "", "", "", ""),
    t("Students", "Frontend", "Students ModuleShell + Masters card grid", "Frontend",
      "Masters cards: Categories, Status, Parent Relationship, Hostel Types, Transport Routes, Groups, Clubs.",
      "Arnav", "Critical", "Phase 1", ""),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Student masters", "DB", "Category, Status, HostelType, TransportRoute, StudentGroup, Club models", "DB",
      "Config masters per functional spec.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Student masters", "Backend/UI", "CRUD APIs + pages for student masters", "All",
      "pages/students/masters/*.",
      "Arnav", "High", "Phase 1", ""),
    t("CORE STUDENT", "", "", "", "", "", "", "", ""),
    t("Student", "DB", "models/Student.js", "DB",
      "student_id, roll_no, application_id?, personal, guardian[], program_id, academic_year_id, status, category_id, documents[].",
      "Arnav", "Critical", "Phase 1", "Single owner — Precious reads via API only."),
    t("Student", "Backend", "CRUD /api/students + profile endpoints", "Backend",
      "Profile update, document management, search/directory.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Student", "Frontend", "StudentList, StudentProfile, DocumentManager", "Frontend",
      "Directory + profile tabs (personal, academic, guardian, docs).",
      "Arnav", "Critical", "Phase 1", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Hostel", "DB/API/UI", "Hostel allocation", "All",
      "Allocate hostel type/room; vacate; report.",
      "Arnav", "Medium", "Phase 2", ""),
    t("Transport", "DB/API/UI", "Transport allocation", "All",
      "Assign route; report.",
      "Arnav", "Medium", "Phase 2", ""),
    t("Medical", "DB/API/UI", "Medical records stub", "All",
      "Basic medical note records on student.",
      "Arnav", "Low", "Phase 3", ""),
    t("Promotion", "Backend/UI", "Semester promotion", "All",
      "Promote based on Academics/Exams results signal; update semester standing.",
      "Arnav", "Critical", "Phase 2", "Consumes result.published / academics promotion contract from Precious."),
    t("Transfer", "Backend/UI", "Student transfer", "All",
      "Transfer program/campus fields; audit.",
      "Arnav", "Medium", "Phase 2", ""),
    t("Alumni", "Backend/UI", "Alumni conversion", "All",
      "On graduation criteria → status Alumni; alumni directory.",
      "Arnav", "High", "Phase 2", ""),
    t("Semester reg handoff", "Backend", "Expose enrolled students for Academics registration", "Backend",
      "GET students by program/semester for Precious Academics.",
      "Arnav", "Critical", "Phase 1", ""),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Directory, Strength, Attendance Summary (read), Hostel, Transport, Alumni", "All",
      "Attendance summary may call Academics attendance API (Precious) read-only.",
      "Arnav", "High", "Phase 2", "Client: student records MIS."),
]

# ---- 06 Faculty ----
sheet_06 = [
    t("MODULE SHELL — FACULTY", "", "", "", "", "", "", "", ""),
    t("Faculty", "Frontend", "Faculty ModuleShell + Masters card grid", "Frontend",
      "Cards: Departments (read-only link), Designations, Faculty Types, Qualifications, Subject Masters, Workload Rules. Match WTI Faculty Master card tone.",
      "Precious", "Critical", "Phase 1", "Do not edit Department model — consume Arnav API."),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Designation", "DB", "models/Designation.js", "DB",
      "name, code, is_active.",
      "Precious", "Critical", "Phase 1", ""),
    t("Faculty Type", "DB", "models/FacultyType.js", "DB",
      "Permanent/Visiting/Trainer etc.",
      "Precious", "High", "Phase 1", ""),
    t("Qualification", "DB", "models/QualificationMaster.js", "DB",
      "degree/cert list for faculty profiles.",
      "Precious", "Medium", "Phase 1", ""),
    t("Subject Master", "DB", "models/Subject.js", "DB",
      "name, code, program_id, theory/practical flag, credits — WTI Subject Master.",
      "Precious", "Critical", "Phase 1", "Arnav Admissions does not own subjects."),
    t("Workload Rules", "DB", "models/WorkloadRule.js", "DB",
      "max_hours_per_week, designation_id?, rules_json.",
      "Precious", "High", "Phase 1", ""),
    t("Masters", "Backend/UI", "CRUD for faculty masters + Subject", "All",
      "academicRoutes; pages/faculty/masters/*.",
      "Precious", "Critical", "Phase 1", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Faculty", "DB", "models/Faculty.js", "DB",
      "employee_code, personal, department_id, designation_id, type_id, qualifications[], subjects[], user_id?, is_active.",
      "Precious", "Critical", "Phase 1", "WTI Faculty Master profiles."),
    t("Faculty", "Backend/UI", "Faculty registration CRUD + directory", "All",
      "Register, map department/designation, link optional User.",
      "Precious", "Critical", "Phase 1", ""),
    t("Subject allocation", "Backend/UI", "Assign subjects to faculty", "All",
      "Respect workload rules + specialization.",
      "Precious", "Critical", "Phase 1", ""),
    t("Timetable assign", "Backend", "API for Academics to attach faculty to slots", "Backend",
      "Expose faculty availability helpers used by Academics timetable.",
      "Precious", "Critical", "Phase 1", "Same owner as Academics — ok to share services."),
    t("Attendance", "Backend/UI", "Faculty attendance / punch or mark", "All",
      "Daily faculty attendance records.",
      "Precious", "High", "Phase 2", "Client: Faculty and attendance reports."),
    t("Leave", "Backend/UI", "Leave management", "All",
      "Apply/approve leave; block timetable conflicts later.",
      "Precious", "High", "Phase 2", ""),
    t("Performance", "Backend/UI", "Performance evaluation stub", "All",
      "Periodic rating records + dashboard widgets.",
      "Precious", "Medium", "Phase 3", ""),
    t("Research/Training", "Backend/UI", "Research & training records", "All",
      "Publications + training log on faculty profile.",
      "Precious", "Low", "Phase 3", ""),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Faculty Directory, Workload, Attendance, Leave, Performance Dashboard", "All",
      "pages/faculty/reports/*.",
      "Precious", "High", "Phase 2", ""),
]

# ---- 07 Academics ----
sheet_07 = [
    t("MODULE SHELL — ACADEMICS", "", "", "", "", "", "", "", ""),
    t("Academics", "Frontend", "Academics ModuleShell + Masters cards", "Frontend",
      "Programs (read), Semesters, Subjects (link), Curriculum, Credits, Academic Calendar, Classrooms, TimeSlots, Batches — WTI Batch/Classroom/Semester parity.",
      "Precious", "Critical", "Phase 1", ""),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Semester", "DB", "models/Semester.js", "DB",
      "name, number, academic_year_id, start/end, is_active — WTI Semester master.",
      "Precious", "Critical", "Phase 1", ""),
    t("Batch", "DB", "models/Batch.js", "DB",
      "name, program_id, academic_year_id, semester_id?, capacity — WTI Batch Master.",
      "Precious", "Critical", "Phase 1", ""),
    t("Classroom", "DB", "models/Classroom.js", "DB",
      "name, type (classroom|lab|workshop), capacity — WTI Classroom master.",
      "Precious", "Critical", "Phase 1", ""),
    t("Time Slot", "DB", "models/TimeSlot.js", "DB",
      "label, start_time, end_time, day_optional.",
      "Precious", "Critical", "Phase 1", ""),
    t("Curriculum", "DB", "models/Curriculum.js + CurriculumSubject", "DB",
      "program_id, semester_id, subject_id, credits, is_elective.",
      "Precious", "Critical", "Phase 1", ""),
    t("Academic Calendar", "DB", "models/AcademicCalendarEvent.js", "DB",
      "events: term start/end, holidays, exam windows.",
      "Precious", "High", "Phase 1", ""),
    t("Masters", "Backend/UI", "CRUD APIs + pages for academics masters", "All",
      "academicRoutes; pages/academics/masters/*.",
      "Precious", "Critical", "Phase 1", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Registration", "Backend/UI", "Semester + subject registration", "All",
      "Register student (from Arnav Students API) into semester subjects.",
      "Precious", "Critical", "Phase 1", "Consume Students API — no Student model edits."),
    t("Batch allocation", "Backend/UI", "Allocate students to batches", "All",
      "Batch membership records.",
      "Precious", "Critical", "Phase 1", ""),
    t("Timetable", "DB", "models/TimetableSlot.js", "DB",
      "batch_id, subject_id, faculty_id, classroom_id, day, time_slot_id.",
      "Precious", "Critical", "Phase 1", ""),
    t("Timetable", "Backend", "AI-assisted conflict-free timetable generator", "Backend",
      "Detect faculty/classroom/batch conflicts; generate draft; publish.",
      "Precious", "Critical", "Phase 2", "AI capability per functional spec."),
    t("Timetable", "Frontend", "Timetable grid UI + publish", "Frontend",
      "Weekly grid; conflict highlights.",
      "Precious", "Critical", "Phase 2", ""),
    t("Attendance", "DB/API/UI", "Student class attendance", "All",
      "Mark per timetable slot; store attendance records.",
      "Precious", "Critical", "Phase 1", "Client faculty/attendance reports feed."),
    t("Assignments/IA", "Backend/UI", "Assignment creation + internal assessments", "All",
      "Internal marks components before Exams module finals.",
      "Precious", "High", "Phase 2", ""),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Timetable, Attendance, Academic Progress, Course Completion, Class Utilization", "All",
      "pages/academics/reports/*.",
      "Precious", "High", "Phase 2", "Client: scheduling + institutional MIS pieces."),
]

# ---- 08 LMS ----
sheet_08 = [
    t("MODULE SHELL — LMS", "", "", "", "", "", "", "", ""),
    t("LMS", "Frontend", "LMS ModuleShell + Masters cards", "Frontend",
      "Course Categories, Learning Paths, Assignment Templates, Quiz Templates, Assessment Types.",
      "Precious", "Critical", "Phase 2", ""),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("LMS masters", "DB", "Category, LearningPath, AssignmentTemplate, QuizTemplate, AssessmentType", "DB",
      "Config masters per spec.",
      "Precious", "High", "Phase 2", ""),
    t("LMS masters", "Backend/UI", "CRUD + master pages", "All",
      "pages/lms/masters/*.",
      "Precious", "High", "Phase 2", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Course", "DB", "models/LmsCourse.js", "DB",
      "title, subject_id?, faculty_id, modules/topics, objectives, published.",
      "Precious", "Critical", "Phase 2", ""),
    t("Content", "Backend/UI", "Content upload + publish", "All",
      "Files/links/videos metadata; publish to enrolled students.",
      "Precious", "Critical", "Phase 2", ""),
    t("Enrollment sync", "Backend", "Enroll students from Academics batch/subject", "Backend",
      "Pull registered students via Academics/Students APIs.",
      "Precious", "Critical", "Phase 2", ""),
    t("Assignment/Quiz", "Backend/UI", "Create assignments & quizzes; student submission", "All",
      "Submission records; due dates.",
      "Precious", "Critical", "Phase 2", ""),
    t("Forums/Live", "Backend/UI", "Discussion forums + live class stub", "All",
      "Forum threads; live class link stub (Zoom/Meet URL).",
      "Precious", "Medium", "Phase 3", ""),
    t("Evaluation", "Backend/UI", "Evaluate submissions + AI-assist stub", "All",
      "Scores + feedback; AI grading assist stub.",
      "Precious", "High", "Phase 2", ""),
    t("Progress", "Backend", "Learning progress + course completion", "Backend",
      "Realtime progress %; completion status.",
      "Precious", "Critical", "Phase 2", ""),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Learning Progress, Course Completion, Assignment Status, Quiz Analytics, Engagement", "All",
      "pages/lms/reports/*.",
      "Precious", "High", "Phase 2", ""),
]

# ---- 09 Exams ----
sheet_09 = [
    t("MODULE SHELL — EXAMS", "", "", "", "", "", "", "", ""),
    t("Exams", "Frontend", "Exams ModuleShell + Masters cards", "Frontend",
      "Exam Types, Grading Scheme, Passing Rules, Evaluation Rules, Result Templates.",
      "Precious", "Critical", "Phase 2", ""),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Exam masters", "DB", "ExamType, GradingScheme, PassingRule, EvaluationRule, ResultTemplate", "DB",
      "Config per functional spec.",
      "Precious", "Critical", "Phase 2", ""),
    t("Exam masters", "Backend/UI", "CRUD + pages", "All",
      "pages/exams/masters/*.",
      "Precious", "Critical", "Phase 2", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Schedule", "DB/API/UI", "Exam scheduling", "All",
      "Dates, subjects, rules; publish to students.",
      "Precious", "Critical", "Phase 2", ""),
    t("Hall ticket", "Backend/UI", "Hall ticket generation", "All",
      "Eligible students; PDF/print stub; center/seating info.",
      "Precious", "Critical", "Phase 2", ""),
    t("Seating", "Backend/UI", "Seating arrangement", "All",
      "Allocate seats/rooms for exam sessions.",
      "Precious", "High", "Phase 2", ""),
    t("Marks", "Backend/UI", "Theory + practical mark entry", "All",
      "Faculty/evaluator entry; lock after submit.",
      "Precious", "Critical", "Phase 2", ""),
    t("Results", "Backend", "Result processing vs grading/passing rules", "Backend",
      "Compute grades; backlog flags.",
      "Precious", "Critical", "Phase 2", ""),
    t("Publish", "Backend/UI", "Publish results + notify student record", "All",
      "Emit result.published for Students/Academics promotion consumers.",
      "Precious", "Critical", "Phase 2", "Contract for Arnav promotion."),
    t("Revaluation", "Backend/UI", "Revaluation workflow", "All",
      "Request → re-enter marks → revise result.",
      "Precious", "Medium", "Phase 3", ""),
    t("Transcript/Certificate", "Backend/UI", "Transcript + certificate generation", "All",
      "PDF stubs; store documents on student via API or local exam docs table referencing student_id.",
      "Precious", "High", "Phase 2", "Client: certification and examination management."),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Result Analysis, Grade Distribution, Pass %, Subject Analysis, Topper, Backlog", "All",
      "pages/exams/reports/*.",
      "Precious", "High", "Phase 2", ""),
]

# ---- 10 Fees ----
sheet_10 = [
    t("MODULE SHELL — FEES", "", "", "", "", "", "", "", ""),
    t("Fees", "Frontend", "Fees ModuleShell + Masters cards", "Frontend",
      "Fee Structure, Fee Heads, Payment Modes, Installments, Scholarships, Fine/Penalty, Tax, Refund Policies.",
      "Arnav", "Critical", "Phase 1", "Client billing→receipt + auto email."),
    t("MASTERS", "", "", "", "", "", "", "", ""),
    t("Fee Head", "DB", "models/FeeHead.js", "DB",
      "name, code, tax_applicable, is_active.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Fee Structure", "DB", "models/FeeStructure.js", "DB",
      "program_id, academic_year_id, heads[], installment_plan_id?, total.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Payment Mode", "DB", "models/PaymentMode.js", "DB",
      "cash, UPI, card, bank, online gateway stub.",
      "Arnav", "High", "Phase 1", ""),
    t("Installment Plan", "DB", "models/InstallmentPlan.js", "DB",
      "schedule JSON: due dates / percentages.",
      "Arnav", "High", "Phase 1", ""),
    t("Fine/Tax/Refund", "DB", "FineRule, TaxConfig, RefundPolicy models", "DB",
      "Per functional spec masters.",
      "Arnav", "High", "Phase 1", ""),
    t("Masters", "Backend/UI", "CRUD + fee master pages", "All",
      "pages/fees/masters/*.",
      "Arnav", "Critical", "Phase 1", ""),
    t("TRANSACTIONS", "", "", "", "", "", "", "", ""),
    t("Assignment", "Backend/UI", "Generate + assign fees to students", "All",
      "Map structure to enrolled students; admission fee assignment.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Collection", "DB", "models/FeePayment.js + StudentFeeLedger", "DB",
      "Payments online/offline; installment tracking; outstanding.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Collection", "Backend/UI", "Collect payment screens + APIs", "All",
      "Record payment; update ledger.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Receipt", "Backend", "Auto receipt generation on successful payment", "Backend",
      "Digital receipt number; PDF stub; store electronic record.",
      "Arnav", "Critical", "Phase 1", "Client process note requirement."),
    t("Tax Invoice", "Backend", "Tax Invoice generation on payment confirmation", "Backend",
      "Generate tax invoice; email invoice + receipt to registered email (mailer).",
      "Arnav", "Critical", "Phase 2", "Client automated communication."),
    t("Scholarship/Fine", "Backend/UI", "Scholarship adjustment + fine calculation", "All",
      "Adjust dues; apply fines per rules.",
      "Arnav", "High", "Phase 2", ""),
    t("Refund", "Backend/UI", "Refund processing", "All",
      "Refund workflow + policy checks; audit.",
      "Arnav", "High", "Phase 2", ""),
    t("Outstanding", "Backend/UI", "Outstanding fee management", "All",
      "Lists, reminders stub, block hall ticket flag API for Exams consumer.",
      "Arnav", "High", "Phase 2", "Precious Exams reads dues flag via API."),
    t("Admission bridge", "Backend", "Admission fee payment confirmation API", "Backend",
      "Endpoint/event used by Admissions enrollment step.",
      "Arnav", "Critical", "Phase 1", "Same owner — internal contract."),
    t("REPORTS", "", "", "", "", "", "", "", ""),
    t("Reports", "All", "Collection, Outstanding, Daily Collection, Student Ledger, Scholarship, Refund, Revenue Dashboard", "All",
      "Also Invoice Status, Receipt Generation, Ageing-style outstanding (client finance-lite).",
      "Arnav", "High", "Phase 2", "Client: Fee collection reports + course-wise profitability later."),
]

# ---- 11 Dashboard ----
sheet_11 = [
    t("INSTITUTION DASHBOARD", "", "", "", "", "", "", "", ""),
    t("Dashboard", "Frontend", "pages/dashboard/InstitutionDashboard.jsx", "Frontend",
      "WTI-style dashboard home: admissions funnel, student strength, fee collection MTD, upcoming exams, attendance snapshot, pending approvals.",
      "Arnav", "Critical", "Phase 2", "Shell/layout already Arnav; widgets call module APIs."),
    t("Dashboard", "Backend", "GET /api/dashboard/summary", "Backend",
      "Aggregates counts Arnav owns (admissions, students, fees); embeds optional precious metrics via HTTP internal calls or shared read models.",
      "Arnav", "Critical", "Phase 2", "Agree metric contracts with Precious."),
    t("Dashboard", "Frontend", "Pending tasks / approvals widget", "Frontend",
      "Pending admission approvals, pending fee follow-ups — process visibility per client note.",
      "Arnav", "High", "Phase 2", ""),
    t("PRECIOUS — MODULE DASHBOARD WIDGETS", "", "", "", "", "", "", "", ""),
    t("Dashboard APIs", "Backend", "GET /api/dashboard/faculty-academics-summary", "Backend",
      "Faculty count, today's classes, attendance %, open exams — academicRoutes.",
      "Precious", "High", "Phase 2", "Arnav dashboard consumes this read API only."),
    t("Dashboard APIs", "Backend", "GET /api/dashboard/lms-exams-summary", "Backend",
      "LMS engagement + exam schedule/result publish stats.",
      "Precious", "Medium", "Phase 2", ""),
    t("CROSS-MODULE MIS", "", "", "", "", "", "", "", ""),
    t("MIS", "Reports", "Institutional MIS report page (admin)", "Frontend",
      "Combined MIS: admissions conversion, fee collection, attendance, pass % — tabs; Arnav page composing APIs.",
      "Arnav", "High", "Phase 3", "Client: Institutional MIS reports."),
    t("MIS", "Reports", "Course-wise profitability (fees vs cost stub)", "All",
      "Revenue by program from Fees; cost stub fields — Phase 3.",
      "Arnav", "Medium", "Phase 3", "Client: Course-wise profitability reports."),
    t("Compliance", "Reports", "Compliance & audit report stub", "All",
      "Export audit log filters; compliance checklist stub.",
      "Arnav", "Medium", "Phase 3", "Client: Compliance and audit reports."),
    t("UI parity", "Frontend", "Dashboard visual match to WTI Dashboard", "Frontend",
      "Spacing, cards, banner, KPI tiles consistent with reference ERP.",
      "Arnav", "High", "Phase 2", "Reference /app/wti/dashboard."),
    t("Module reports registry", "Docs", "List each module report owner + path", "Docs",
      "Ensure every report in functional spec is tracked Done in module sheets.",
      "Both", "Medium", "Phase 2", ""),
]

# ---- 12 Phases ----
sheet_12 = [
    t("PHASE MILESTONES", "", "", "", "", "", "", "", ""),
    t("Phase 0", "Milestone", "Scaffold + merge-safe routing + ERP shell + RBAC + login merged", "All",
      "Precious can branch; nav shows 7 modules (empty shells ok).",
      "Both", "Critical", "Phase 0", "Arnav merges first."),
    t("Phase 0", "UI", "Left nav labels exactly: Admissions, Students, Faculty, Academics, LMS, Exams, Fees", "Frontend",
      "Order and naming locked.",
      "Arnav", "Critical", "Phase 0", ""),
    t("Phase 0", "UI", "ModuleShell tabs Dashboard | Masters | Transactions | Reports on each module stub", "Frontend",
      "Empty states with WTI-like banner.",
      "Arnav", "Critical", "Phase 0", "Precious uses same shell for owned modules."),
    t("Phase 1", "Milestone", "Shared masters + Admissions + Students + Fees core + Faculty masters/profiles + Academics masters/registration/attendance", "All",
      "End-to-end: apply → pay → enroll → appear in academics registration.",
      "Both", "Critical", "Phase 1", ""),
    t("Phase 1", "Contract test", "admission.enrolled creates Student readable by Precious", "All",
      "Integration smoke test documented.",
      "Both", "Critical", "Phase 1", ""),
    t("Phase 2", "Milestone", "Timetable AI, LMS core, Exams core, Fee tax invoice email, Dashboards, major reports", "All",
      "Teaching/learning/exam/fee loop working.",
      "Both", "Critical", "Phase 2", ""),
    t("Phase 2", "Contract test", "result.published usable for promotion; dues flag blocks hall ticket", "All",
      "Cross-module smoke tests.",
      "Both", "Critical", "Phase 2", ""),
    t("Phase 3", "Milestone", "Alumni, revaluation, forums/live polish, performance, profitability/compliance MIS, AI OCR real integration optional", "All",
      "Polish + advanced items.",
      "Both", "High", "Phase 3", ""),
    t("UI PARITY CHECKLIST (vs Celeris WTI)", "", "", "", "", "", "", "", ""),
    t("UI", "Parity", "Left sidebar module list styling", "Frontend",
      "Active state, icons, collapsible behavior similar to reference.",
      "Arnav", "High", "Phase 0", "https://erp.smart-aiapps.com/app/wti/dashboard"),
    t("UI", "Parity", "Top context bar (workspace, search, FY, location)", "Frontend",
      "Present on all pages.",
      "Arnav", "High", "Phase 0", ""),
    t("UI", "Parity", "Blue module banner title + subtitle + badge count on Masters", "Frontend",
      "Masters landing shows module count badge like '11 MODULES'.",
      "Arnav", "High", "Phase 1", ""),
    t("UI", "Parity", "Masters card grid: icon, title, description, Open module link", "Frontend",
      "Same interaction pattern as WTI Masters.",
      "Arnav", "Critical", "Phase 1", "Precious follows same card component."),
    t("UI", "Parity", "List/table pages consistent density and primary actions", "Frontend",
      "Shared DataTable.",
      "Both", "High", "Phase 1", ""),
    t("UAT / ACCEPTANCE", "", "", "", "", "", "", "", ""),
    t("UAT", "Admissions", "Full workflow application→enrollment with fee payment", "All",
      "Demo script pass.",
      "Arnav", "Critical", "Phase 1", ""),
    t("UAT", "Students", "Profile, hostel/transport, promotion, alumni path", "All",
      "Demo script pass.",
      "Arnav", "High", "Phase 2", ""),
    t("UAT", "Faculty/Academics", "Faculty assign → timetable → attendance", "All",
      "Demo script pass.",
      "Precious", "Critical", "Phase 2", ""),
    t("UAT", "LMS/Exams", "Course content → submission → exam → result publish", "All",
      "Demo script pass.",
      "Precious", "Critical", "Phase 2", ""),
    t("UAT", "Fees", "Assign, collect, receipt, email invoice/receipt, outstanding", "All",
      "Demo script pass per client process note.",
      "Arnav", "Critical", "Phase 2", ""),
    t("UAT", "E2E", "Single student journey application to alumni (happy path)", "All",
      "No duplicate data entry across modules.",
      "Both", "Critical", "Phase 3", ""),
    t("UAT", "RBAC", "Each role sees only permitted nav modules/actions", "All",
      "Permission matrix verified.",
      "Arnav", "Critical", "Phase 1", ""),
    t("Docs", "Handoff", "README + seed credentials + tracker status updated", "Docs",
      "Keep this tracker Status column current as work proceeds.",
      "Both", "High", "Phase 0", ""),
]

sheets = {
    "00 Overview": None,  # special
    "01 Architecture-Foundation": sheet_01,
    "02 RBAC-Auth-Users": sheet_02,
    "03 Shared-Platform-Masters": sheet_03,
    "04 Admissions": sheet_04,
    "05 Students": sheet_05,
    "06 Faculty": sheet_06,
    "07 Academics": sheet_07,
    "08 LMS": sheet_08,
    "09 Exams": sheet_09,
    "10 Fees": sheet_10,
    "11 Dashboard-Reports": sheet_11,
    "12 Phases-Acceptance": sheet_12,
}


def style_header(ws):
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, 11):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws):
    widths = [28, 18, 52, 12, 70, 14, 12, 12, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def count_tasks(rows):
    return sum(1 for r in rows if r[2] and r[3])  # has Item and Layer


def write_task_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append(COLS)
    style_header(ws)
    section_fill = PatternFill("solid", fgColor="DBEAFE")
    for r in rows:
        ws.append(r)
        if r[0] and not r[2]:  # section banner row
            for c in range(1, 11):
                ws.cell(ws.max_row, c).fill = section_fill
                ws.cell(ws.max_row, c).font = Font(bold=True)
    autosize(ws)
    ws.freeze_panes = "A2"
    return count_tasks(rows)


def main():
    wb = Workbook()
    # Overview
    ws = wb.active
    ws.title = "00 Overview"
    for row in overview_header_rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14, color="1D4ED8")
    counts = {}
    # temporary counts from data
    data_counts = {
        "01 Architecture-Foundation": count_tasks(sheet_01),
        "02 RBAC-Auth-Users": count_tasks(sheet_02),
        "03 Shared-Platform-Masters": count_tasks(sheet_03),
        "04 Admissions": count_tasks(sheet_04),
        "05 Students": count_tasks(sheet_05),
        "06 Faculty": count_tasks(sheet_06),
        "07 Academics": count_tasks(sheet_07),
        "08 LMS": count_tasks(sheet_08),
        "09 Exams": count_tasks(sheet_09),
        "10 Fees": count_tasks(sheet_10),
        "11 Dashboard-Reports": count_tasks(sheet_11),
        "12 Phases-Acceptance": count_tasks(sheet_12),
    }
    for name, scope, _old, owner in overview_sheets_index:
        ws.append([name, scope, data_counts[name], owner, "", "", "", "", "", ""])
    total = sum(data_counts.values())
    ws.append([])
    ws.append(["TOTAL TASKS", total, "", "", "", "", "", "", "", ""])
    ws["A" + str(ws.max_row)].font = Font(bold=True)
    ws.append([])
    for r in merge_rules:
        ws.append(r)
        if r[0].startswith("MERGE"):
            ws.cell(ws.max_row, 1).font = Font(bold=True, color="B91C1C")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 88
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 48

    for title, rows in sheets.items():
        if rows is None:
            continue
        write_task_sheet(wb, title, rows)

    xlsx_path = os.path.join(OUT_DIR, "Smart_Institute_AI_Implementation_Tracker.xlsx")
    wb.save(xlsx_path)

    # Also write a combined CSV + per-sheet CSVs folder
    csv_dir = os.path.join(OUT_DIR, "implementation_tracker_csv")
    os.makedirs(csv_dir, exist_ok=True)

    # Overview as simple csv
    with open(os.path.join(csv_dir, "00_Overview.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Sheet", "Scope", "Item Count", "Primary Owner"])
        for name, scope, _o, owner in overview_sheets_index:
            w.writerow([name, scope, data_counts[name], owner])
        w.writerow([])
        w.writerow(["TOTAL TASKS", total])
        w.writerow([])
        w.writerow(["MERGE-CONFLICT RULES"])
        for r in merge_rules[1:]:
            if r[0]:
                w.writerow([r[0]])

    all_rows = []
    for title, rows in sheets.items():
        if rows is None:
            continue
        safe = title.replace(" ", "_")
        path = os.path.join(csv_dir, f"{safe}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(COLS)
            for r in rows:
                w.writerow(r)
                if r[2] and r[3]:
                    all_rows.append([title] + r)

    with open(os.path.join(OUT_DIR, "Smart_Institute_AI_Implementation_Tracker_ALL.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Sheet"] + COLS)
        for r in all_rows:
            w.writerow(r)

    # Ownership summary
    arnav = sum(1 for r in all_rows if r[7] == "Arnav")
    precious = sum(1 for r in all_rows if r[7] == "Precious")
    both = sum(1 for r in all_rows if r[7] == "Both")
    print(f"Saved: {xlsx_path}")
    print(f"CSV folder: {csv_dir}")
    print(f"Combined CSV: Smart_Institute_AI_Implementation_Tracker_ALL.csv")
    print(f"TOTAL TASKS: {total}")
    print(f"Arnav: {arnav} | Precious: {precious} | Both: {both}")


if __name__ == "__main__":
    main()
